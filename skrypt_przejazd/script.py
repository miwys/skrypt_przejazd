# © openrouteservice.org by HeiGIT | Map data © OpenStreetMap contributors

import requests
import json
import openpyxl as xl
import pandas as pd

def load_data():
    wb = xl.load_workbook('input.xlsx') #plik z którego pobierane są dane
    sheet = wb.active
    coord = []
    sources = []
    destinations = []

    for i in range(1,3): 
        for row in range(2, sheet.max_row+1): # w funkcju range(arg1 - wiersz od którego mają wczytywać się dane, arg2 - liczba wierszy)  
            cell = sheet.cell(row,i) # pobranie danych z komórki (arg1 - nr. wiersza, arg2 -  nr. kolumny)
            if cell.value is None:
                break 
            spl = str(cell.value).split(',')
            lon = float(spl[1])
            lat = float(spl[0])
            coord.append([lon,lat])
        if i == 1:
            destinations = coord
            coord = []            
        elif i == 2:
            sources = coord
    return destinations, sources

dest_all, src_all = load_data()

src = []
for i in range(len(src_all)):
    src.append(i+1)

n = 300 #liczba 
final = [src[i * n:(i + 1) * n] for i in range((len(src) + n - 1) // n )] #podzielenie listy indeksów punktów początkowych na n-elementowe podzbiory. Dla większych wartości n wykorzystywany serwis może błędnie działać

df_dur = pd.DataFrame(range(len(src_all)))
df_dist = pd.DataFrame(range(len(src_all)))

for i in range(len(dest_all)):
    loc = src_all.copy()
    loc.insert(0, dest_all[i])
    x = []
    y = []

    for j in range(len(final)):
        body = {"locations":loc,"destinations":[0],"metrics":["distance","duration"],"sources":final[j],"units":"m"}
        #locations: lista wszystkich wykorzystywanych lokalizacji
        #destinations: lista indeksów lokalizacji wykorzystanej jako punkt końcowy
        #sources: lista indeksów lokalizacji wykorzystywanej jako punkt początkowy
        headers = {
            'Accept': 'application/json, application/geo+json, application/gpx+xml, img/png; charset=utf-8',
            'Authorization': '5b3ce3597851110001cf624856cbbda85f7a4c688fa9f852ab15eded', #klucz autoryzacji 
            'Content-Type': 'application/json; charset=utf-8'
        }
        call = requests.post('https://api.openrouteservice.org/v2/matrix/driving-car', json=body, headers=headers)
        print(call.status_code, call.reason, f'step: {i}/{len(dest_all)}')
        data = json.loads(call.text)
        x = x + data['durations']
        y = y + data['distances']   

    df_dur[i] = pd.DataFrame(x)
    df_dist[i] = pd.DataFrame(y)


with pd.ExcelWriter('output.xlsx') as writer:
        df_dur.to_excel(writer, sheet_name='czas')
        df_dist.to_excel(writer, sheet_name='odległość')


